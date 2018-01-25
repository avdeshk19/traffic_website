# dsvjsbvisbvsjvb
# import json
# from urllib.request import urlopen
# import requests
# import math
# from xlutils.copy import copy    
# from xlrd import open_workbook
# import pandas as pd
# import openpyxl
# import datetime
# import os
# import pathlib
import datetime
# from datetime import timedelta
import json
from openpyxl import load_workbook

from flask import Flask, render_template, request, json,jsonify
app = Flask(__name__, template_folder = 'template', static_url_path = '/static')

wb = load_workbook('./assets/nodes.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
lat = []
longi = []
loc = []
dist = []
for i in range(2, 15):
	lat.append(sheet['A'+str(i)].value)
	longi.append(sheet['B'+str(i)].value)
	loc.append(sheet['E'+str(i)].value)
	if(i!=14):
		dist.append(sheet['D'+str(i)].value)
api_down = []
for i in range(0,14):
	time = []
	for j in range(0,24):
		time.append(0)
	api_down.append(time)
wb = load_workbook('./assets/downroute_api.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
for i in range(2, 314):
	time_val = int(sheet['D'+str(i)].value)
	speed_val = sheet['C'+str(i)].value
	node = int(sheet['B'+str(i)].value)
	api_down[node][time_val] = speed_val

api_up = []
for i in range(0,14):
	time = []
	for j in range(0,24):
		time.append(0)
	api_up.append(time)
wb = load_workbook('./assets/uproute_api.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
for i in range(2, 314):
	time_val = int(sheet['D'+str(i)].value)
	speed_val = sheet['C'+str(i)].value
	node = int(sheet['B'+str(i)].value)
	api_up[node][time_val] = speed_val
# sheet['A1'].value

# workbook1 = xlrd.open_workbook('nodes.xlsx', on_demand=True)
# sheet1 = workbook1.sheet_by_index(0)
# arrayofnodelat = sheet1.col_values(0)
# arrayofnodelong = sheet1.col_values(1)
# dist = sheet1.col_values(3)
# loc = sheet1.col_values(4)

# Load in the workbook


# config
# app.secret_key = 'my precious'
@app.route('/traffic', methods = ['GET'])
def home():
	# for i in loc:
	# 	print (i)
	# print("hello")
	# colours = ['Uproute1', Upr]
    # return render_template('test.html', colours=colours)
	return render_template('traffic.html')
	# return "Hello, World!"
	
@app.route('/traffic/map', methods = ['GET','POST'])
def read_data():
	# print("here")
	if request.method=='POST':
		date = request.form['date']
		# date = request.form('something')
		uproute = request.form['uproute']
		downroute = request.form['downroute']
		wb = load_workbook('./assets/'+str(date)+'.xlsx')
		sheet = wb.active
		i = 2
		j = i+1
		flag = 0

		if(uproute=='0' and downroute!='0'): # downroute val = 0
			max_rows = sheet.max_row
			# print(max_rows)
			# print("hello")
			route_no= 0
			flag = 0
			while(j<max_rows-1):
				if(sheet['C'+str(i)].value>sheet['C'+str(j)].value):
					route_no = route_no+1
					if(str(route_no)==downroute):
						break
					else:
						while(j<max_rows-1 and sheet['C'+str(i)].value>sheet['C'+str(j)].value):
							i = i+1
							j = i+1
				i = i+1
				j = i+1

			if(j==max_rows-1):
				flag = 1
			if(flag!=1):
				start =i
				while(j < max_rows - 1 and sheet['C'+str(i)].value>sheet['C'+str(j)].value):
					i = i +1
					j=i+1
				end = j
				data = []
				for i in range(start,end):
					l=dict()
					j = i+1
					node = int(sheet['C'+str(i)].value)
					# l.append(node)
					l["node"] = node
					l["latitude"] = float(lat[node-1])
					l["longitude"] = float(longi[node-1])
					l["location"] = str(loc[node-1])
					l["up_or_down"]=0
					# l.append(float(lat[node-1]))
					# l.append(float(longi[node-1]))
					# l.append(loc[node-1])
					speed = float(sheet['D'+str(i)].value)
					l["speed_bus_data"] = speed
					# l.append(speed)
					tim1 = sheet['E'+str(i)].value
					str1 = tim1
					idx2 =-1
					idx = -1
					for k in range(0,len(str1)):
						if(str1[k]==','):
							idx = k+2

						if(str1[k]=='.'):
							idx2=k
							break

					str2 = str1[0:idx2]
					hr = str1[idx]+str1[idx+1]
					mn = str1[idx+3]+str1[idx+4]
					l["timet"]=str2;
					if(i!=start):
						hr = int(hr)
						mn = int(mn)
						curr_time = datetime.datetime(2000, 2, 1, hr, mn)
						prev_time = datetime.datetime(2000, 2, 1, lhr, lmn)
						diff = curr_time-prev_time
						diff, sec = divmod(diff.total_seconds(), 60)
						l["diff"] = diff
						if(node<13):
							l["dist"] = dist[node-1]
							# l.append(dist[node-1])
					final1 = str1[idx] + str1[idx+1]
					idx = int(final1)
					l["speed_google"] = api_down[node][idx]
					# print(idx)
					# l.append(api_down[node][idx])
					# l = str(l)
					data.append(l)
					# data[node] = l
					lhr = int(hr)
					lmn = int(mn)
					# print(node, " ", l)
					# jsonStr = json.dumps(data)
				data = str(data)	
				jsonStr = json.dumps(data)
				# print(jsonStr)
				# return jsonify(Employees=jsonStr)
				# data = '{"firstname": "Mr.", "lastname": "My Father\u0027s Son"}'
				final = ""
				for c in data:
					if(c == '\''):
						final += '\"'
					else:
						final += c

				# print(final)
				return render_template('traffic.html', result  = final)        


		elif(uproute!='0' and downroute=='0'): #uproute val = 1
			max_rows = sheet.max_row
			# print(max_rows)
			# print("hello")
			route_no= 0
			flag = 0
			while(j<max_rows-1):
				if(sheet['C'+str(i)].value<sheet['C'+str(j)].value):
					route_no = route_no+1
					if(str(route_no)==uproute):
						break
					else:
						while(j<max_rows-1 and sheet['C'+str(i)].value<sheet['C'+str(j)].value):
							i = i+1
							j = i+1
				i = i+1
				j = i+1

			if(j==max_rows-1):
				flag = 1
			if(flag!=1):
				start =i
				while(j < max_rows - 1 and sheet['C'+str(i)].value<sheet['C'+str(j)].value):
					i = i +1
					j=i+1
				end = j
				data = []
				for i in range(start,end):
					l=dict()
					j = i+1
					node = int(sheet['C'+str(i)].value)
					# l.append(node)
					l["node"] = node
					l["latitude"] = float(lat[node-1])
					l["longitude"] = float(longi[node-1])
					l["location"] = str(loc[node-1])

					# l.append(float(lat[node-1]))
					# l.append(float(longi[node-1]))
					# l.append(loc[node-1])
					speed = float(sheet['D'+str(i)].value)
					l["speed_bus_data"] = speed
					# l.append(speed)
					tim1 = sheet['E'+str(i)].value
					str1 = tim1
					idx2 =-1
					idx = -1
					for k in range(0,len(str1)):
						if(str1[k]==','):
							idx = k+2

						if(str1[k]=='.'):
							idx2=k
							break

					str2 = str1[0:idx2]
					hr = str1[idx]+str1[idx+1]
					mn = str1[idx+3]+str1[idx+4]
					l["timet"]=str2;
					if(i!=start):
						hr = int(hr)
						mn = int(mn)
						curr_time = datetime.datetime(2000, 2, 1, hr, mn)
						prev_time = datetime.datetime(2000, 2, 1, lhr, lmn)
						diff = curr_time-prev_time
						diff, sec = divmod(diff.total_seconds(), 60)
						l["diff"] = diff
						if(node<13):
							l["dist"] = dist[node-1]
							# l.append(dist[node-1])
					else:
						l["dist"] = dist[node-1]
					final1 = str1[idx] + str1[idx+1]
					idx = int(final1)
					l["speed_google"] = api_up[node][idx]
					# print(idx)
					l["up_or_down"]=1
					# l.append(api_down[node][idx])
					# l = str(l)
					data.append(l)
					# data[node] = l
					lhr = int(hr)
					lmn = int(mn)
					# print(node, " ", l)
					# jsonStr = json.dumps(data)
				data = str(data)	
				jsonStr = json.dumps(data)
				# print(jsonStr)
				# return jsonify(Employees=jsonStr)
				# data = '{"firstname": "Mr.", "lastname": "My Father\u0027s Son"}'
				final = ""
				for c in data:
					if(c == '\''):
						final += '\"'
					else:
						final += c

				# print(final)
				return render_template('traffic.html', result  = final)	# while(i<max_rows)


		# print(date)
		# print(uproute)
		# print(downroute)
		# return (date)
		# if(uproute==1):

	return render_template('traffic.html')
	
    # if request.method=='GET':
	
    
    	

if __name__=='__main__':
    app.secret_key='super_secret_key'
    app.debug=True
    app.run(host='0.0.0.0',port=5000)